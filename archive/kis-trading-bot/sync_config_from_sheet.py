#!/usr/bin/env python3
"""
Google Sheet(Account/Key) -> local kis-demo-config.json 동기화
필수: OAuth 토큰(token.json) + client_secret.json
"""

import json
import os
from pathlib import Path
from typing import Tuple

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
from openpyxl import load_workbook

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
DEFAULT_SHEET_ID = "1NEgJIigfVQuuS4Ox9uldLGCrgVmZ79ro64FkQYTsli0"  # AItrading-demo


def get_creds(client_secret_path: Path, token_path: Path) -> Credentials:
    creds = None
    client = json.loads(client_secret_path.read_text(encoding='utf-8')).get('installed', {})

    if token_path.exists():
        try:
            # 표준 authorized_user 형식
            creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
        except Exception:
            # node/googleapis 형식(access_token/refresh_token만 있는 파일) 호환
            raw = json.loads(token_path.read_text(encoding='utf-8'))
            if raw.get('refresh_token'):
                creds = Credentials(
                    token=raw.get('access_token'),
                    refresh_token=raw.get('refresh_token'),
                    token_uri='https://oauth2.googleapis.com/token',
                    client_id=client.get('client_id'),
                    client_secret=client.get('client_secret'),
                    scopes=SCOPES,
                )

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(client_secret_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")

    return creds


def export_sheet_xlsx(drive, file_id: str, out_path: Path):
    request = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    out_path.write_bytes(fh.getvalue())


def read_key_account(xlsx_path: Path) -> Tuple[str, str, str]:
    wb = load_workbook(str(xlsx_path), data_only=True)
    key_ws = wb["Key"]
    acc_ws = wb["Account"]

    kis_key = None
    kis_secret = None
    account_no = None

    for row in key_ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=3, values_only=True):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() == "KIS_KEY" or str(row[1]).strip() == "KIS_KEY":
            kis_key = str(row[1] if str(row[0]).strip() == "KIS_KEY" else row[2]).strip()
        if str(row[0]).strip() == "KIS_SECRET" or str(row[1]).strip() == "KIS_SECRET":
            kis_secret = str(row[1] if str(row[0]).strip() == "KIS_SECRET" else row[2]).strip()

    for row in acc_ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=3, values_only=True):
        if not row or row[0] is None:
            continue
        if str(row[0]).strip() == "계좌번호" or str(row[1]).strip() == "계좌번호":
            account_no = str(row[1] if str(row[0]).strip() == "계좌번호" else row[2]).strip()

    if not (kis_key and kis_secret and account_no):
        raise RuntimeError("시트에서 KIS_KEY/KIS_SECRET/계좌번호를 찾지 못했습니다.")

    return kis_key, kis_secret, account_no


def main():
    base = Path(__file__).resolve().parent

    client_secret = Path(os.getenv("GOOGLE_CLIENT_SECRET_PATH", str(base.parent / "client_secret.json")))
    token_path = Path(os.getenv("GOOGLE_TOKEN_PATH", str(base.parent / "token.json")))
    sheet_id = os.getenv("GOOGLE_SHEET_ID", DEFAULT_SHEET_ID)

    if not client_secret.exists():
        raise FileNotFoundError(f"client secret not found: {client_secret}")

    creds = get_creds(client_secret, token_path)
    drive = build("drive", "v3", credentials=creds)

    xlsx_path = base / "_sheet_tmp.xlsx"
    export_sheet_xlsx(drive, sheet_id, xlsx_path)

    kis_key, kis_secret, account_no = read_key_account(xlsx_path)

    config = {
        "KIS_KEY": kis_key,
        "KIS_SECRET": kis_secret,
        "계좌번호": account_no,
    }

    out_path = base.parent / "kis-demo-config.json"
    out_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    try:
        xlsx_path.unlink(missing_ok=True)
    except Exception:
        pass

    print(f"✅ config synced: {out_path}")


if __name__ == "__main__":
    main()
